"""
Medal Showcase - Excel coordinate parsing and showcase data builder.

Parses IL-2_Tracker_award_coordinates.xlsx to build per-country medal
placement data for the Medal Showcase modal.

Excel format (Sheet1):
  - Section header rows in column A:  "USSR (early)", "USSR (late)", "USA",
    "Germany", "Britain"
  - Within each section:
      Col A: medal name (no extension, e.g. "iron_cross_2nd_big1") or
             overlay identifier (e.g. "Germany_overlay")
      Col B: x  (top-left, pixels)
      Col C: y  (top-left, pixels)
      Col D: width
      Col E: height
      Col F: unused
      Col G-K: replacement names (higher-tier variants, left=base+1 …
               right=highest)

Country keys used throughout:
  "ussr_early", "ussr_late", "usa", "germany", "britain"

Asset folder mapping (relative to CampaignRanksAwards/):
  ussr_early  -> USSR/early
  ussr_late   -> USSR/late
  usa         -> US
  germany     -> Germany
  britain     -> Britain

Developer notes
---------------
* Excel is the single source of truth for coordinates.  Never hard-code
  positions in Python or JavaScript.
* Overlay rows are identified by the word "overlay" in their name
  (case-insensitive) or the special marker "glass_overlay_ussr".
* Replacement selection: for each slot (column A), collect the chain
  [col_A] + [col_G … col_K].  Walk from highest (rightmost) to lowest
  (col_A).  Show the first entry in that chain that the player has earned.
* USSR early/late decision: compare the campaign's last mission date against
  USSR_REDESIGN_DATE.  Awards earned before that date use early assets;
  awards earned on/after use late assets.  For the showcase we use whichever
  canvas matches the campaign's current era.
* The "no re-earning" rule (USSR) is enforced by max_awards:1 in the award
  config, so no extra logic is needed here.
"""

from __future__ import annotations

import logging
from datetime import date, datetime
from functools import lru_cache
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Import the canonical redesign date from path_utils (same constant used by
# resolve_country_folder for step3 asset selection).
try:
    from campaign_service_record.utils.path_utils import USSR_TRANSITION_DATE as USSR_REDESIGN_DATE
except ImportError:
    USSR_REDESIGN_DATE = date(1943, 1, 6)  # fallback

# Section header in Excel  ->  internal country key
SECTION_HEADERS: dict[str, str] = {
    'USSR (early)': 'ussr_early',
    'USSR (late)': 'ussr_late',
    'USA': 'usa',
    'Germany': 'germany',
    'Britain': 'britain',
}

# internal country key  ->  relative folder under CampaignRanksAwards/
ASSET_FOLDER: dict[str, str] = {
    'ussr_early': 'USSR/early',
    'ussr_late':  'USSR/late',
    'usa':        'US',
    'germany':    'Germany',
    'britain':    'Britain',
}

# country key  ->  canvas filename (just the filename, not full path)
CANVAS_FILENAME: dict[str, str] = {
    'ussr_early': 'USSR_canvas.png',
    'ussr_late':  'USSR_canvas.png',
    'usa':        'usa_canvas.png',
    'germany':    'Germany_canvas.png',
    'britain':    'Britain_canvas.png',
}

# country key  ->  overlay filename
OVERLAY_FILENAME: dict[str, str] = {
    'ussr_early': 'USSR_overlay.png',
    'ussr_late':  'USSR_overlay.png',
    'usa':        'usa_overlay.png',
    'germany':    'Germany_overlay.png',
    'britain':    'Britain_overlay.png',
}

# Normalise "Soviet Union" / "USSR" from campaign data to internal key
_CAMPAIGN_COUNTRY_MAP: dict[str, str] = {
    'germany':             'germany',
    'britain':             'britain',
    'uk':                  'britain',
    'great britain':       'britain',
    'united kingdom':      'britain',
    'usa':                 'usa',
    'us':                  'usa',
    'united states':       'usa',
    'united states of america': 'usa',
    'soviet union':        'ussr',
    'ussr':                'ussr',
    'russia':              'ussr',
}


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

class MedalEntry:
    """One medal slot from the Excel, with optional replacement chain."""
    __slots__ = ('name', 'x', 'y', 'w', 'h', 'replacements')

    def __init__(
        self,
        name: str,
        x: int, y: int, w: int, h: int,
        replacements: list[str],
    ) -> None:
        self.name         = name
        self.x            = x
        self.y            = y
        self.w            = w
        self.h            = h
        self.replacements = replacements  # ordered from base+1 to highest

    def full_chain(self) -> list[str]:
        """Return [base, *replacements] – all candidates, lowest first."""
        return [self.name] + self.replacements

    def to_dict(self) -> dict:
        return {
            'name':         self.name,
            'x':            self.x,
            'y':            self.y,
            'w':            self.w,
            'h':            self.h,
            'replacements': self.replacements,
        }


class CountryShowcase:
    """All placement data for one country key."""
    __slots__ = ('medals', 'overlay')

    def __init__(
        self,
        medals: list[MedalEntry],
        overlay: Optional[MedalEntry],
    ) -> None:
        self.medals  = medals
        self.overlay = overlay


# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------

def _is_overlay_name(name: str) -> bool:
    """Return True if this row is the overlay entry, not a medal."""
    low = name.lower()
    return 'overlay' in low or low == 'glass_overlay_ussr'


def _try_fix_name(name: str) -> str:
    """
    Apply known normalisation fixes.

    Examples:
      "AF_cross_2bars_big"   -> "AF_cross_2bars_big1"   (missing trailing 1)
      "order_surov_big1"     -> kept as-is (typo in Excel for order_suvorov)
    """
    if name.endswith('_big') and not name.endswith('_big1'):
        fixed = name + '1'
        logger.warning(
            "[medal_showcase] Excel name '%s' appears to be missing '1' suffix"
            " – treating as '%s'", name, fixed)
        return fixed
    return name


def parse_excel_coordinates(excel_path: Path) -> dict[str, CountryShowcase]:
    """
    Parse the award-coordinates Excel file into a per-country dict.

    Returns:
        { country_key: CountryShowcase }

    Raises:
        ImportError  if openpyxl is unavailable
        FileNotFoundError / ValueError  on file / sheet issues
    """
    try:
        import openpyxl  # noqa: PLC0415
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required to parse medal coordinates.  "
            "Install it with: pip install openpyxl"
        ) from exc

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    wb = openpyxl.load_workbook(str(excel_path), data_only=True, read_only=True)

    if 'Sheet1' not in wb.sheetnames:
        raise ValueError(
            f"Expected sheet 'Sheet1' in {excel_path}.  "
            f"Available sheets: {wb.sheetnames}"
        )

    ws = wb['Sheet1']

    result: dict[str, CountryShowcase] = {}
    current_key: Optional[str] = None
    medals_buf: list[MedalEntry] = []
    overlay_buf: Optional[MedalEntry] = None

    def _flush() -> None:
        nonlocal medals_buf, overlay_buf
        if current_key is not None:
            result[current_key] = CountryShowcase(
                medals=list(medals_buf),
                overlay=overlay_buf,
            )
        medals_buf  = []
        overlay_buf = None

    for raw_row in ws.iter_rows(values_only=True):
        col_a = raw_row[0]
        if col_a is None:
            continue

        col_a_str = str(col_a).strip()
        if not col_a_str:
            continue

        # --- Section header? ---
        if col_a_str in SECTION_HEADERS:
            _flush()
            current_key = SECTION_HEADERS[col_a_str]
            medals_buf  = []
            overlay_buf = None
            continue

        if current_key is None:
            continue  # rows before first section header

        # --- Coordinate row ---
        try:
            x_raw = raw_row[1]
            y_raw = raw_row[2]
            w_raw = raw_row[3]
            h_raw = raw_row[4]

            if any(v is None for v in (x_raw, y_raw, w_raw, h_raw)):
                continue  # header sub-row ("x y w h") or empty

            x = int(x_raw)
            y = int(y_raw)
            w = int(w_raw)
            h = int(h_raw)
        except (TypeError, ValueError):
            continue

        # Replacements: columns G–K  (indices 6–10)
        replacements: list[str] = []
        for idx in range(6, 11):
            raw_val = raw_row[idx] if idx < len(raw_row) else None
            if raw_val is not None:
                rep = str(raw_val).strip()
                if rep:
                    rep = _try_fix_name(rep)
                    replacements.append(rep)

        name = _try_fix_name(col_a_str)

        entry = MedalEntry(name, x, y, w, h, replacements)

        if _is_overlay_name(name):
            overlay_buf = entry
        else:
            medals_buf.append(entry)

    # Flush the last section
    _flush()

    logger.info(
        "[medal_showcase] Parsed %d country sections from %s",
        len(result), excel_path.name,
    )
    return result


# ---------------------------------------------------------------------------
# Cached loader
# ---------------------------------------------------------------------------

_cached_coordinates: Optional[dict[str, CountryShowcase]] = None
_cached_excel_mtime: Optional[float] = None


def load_coordinates(excel_path: Path) -> dict[str, CountryShowcase]:
    """
    Return parsed coordinates, re-parsing if the Excel file has changed.
    Thread-safe via GIL (single writer, dict swap).
    """
    global _cached_coordinates, _cached_excel_mtime

    try:
        mtime = excel_path.stat().st_mtime
    except OSError:
        mtime = None

    if _cached_coordinates is None or mtime != _cached_excel_mtime:
        _cached_coordinates = parse_excel_coordinates(excel_path)
        _cached_excel_mtime = mtime

    return _cached_coordinates


# ---------------------------------------------------------------------------
# Country resolution helpers
# ---------------------------------------------------------------------------

def resolve_showcase_country(campaign_country: str) -> str:
    """
    Map a campaign's country string to an internal country key.

    For USSR, returns "ussr" (caller must decide early/late based on date).
    """
    return _CAMPAIGN_COUNTRY_MAP.get((campaign_country or '').strip().lower(), '')


def resolve_ussr_variant(last_mission_date: Optional[str]) -> str:
    """
    Decide "ussr_early" or "ussr_late" based on the campaign's last mission.
    Falls back to "ussr_early" if date cannot be parsed.
    """
    if not last_mission_date:
        return 'ussr_early'
    for fmt in ('%Y-%m-%d', '%d.%m.%Y'):
        try:
            d = datetime.strptime(last_mission_date, fmt).date()
            return 'ussr_late' if d >= USSR_REDESIGN_DATE else 'ussr_early'
        except ValueError:
            continue
    return 'ussr_early'


# ---------------------------------------------------------------------------
# Award-name  <->  showcase-name mapping
# ---------------------------------------------------------------------------

def award_image_to_showcase_name(image_field: str) -> str:
    """
    Derive the showcase identifier (without extension) from the award's
    image field stored in campaign_events.json.

    Examples:
        "iron_cross_2nd.dds"               -> "iron_cross_2nd_big1"
        "CampaignRanksAwards/Germany/dfc.dds" -> "dfc_big1"
        "dfc_bar.dds"                      -> "dfc_bar_big1"
        "air_medal_big.dds"                -> "air_medal_big1"
    """
    if not image_field:
        return ''
    stem = Path(image_field).stem          # strip path + extension
    if stem.endswith('_big'):
        return stem + '1'
    return stem + '_big1'


def _resolve_asset_file(
    showcase_name: str,
    country_key: str,
    assets_dir: Path,
) -> Optional[Path]:
    """
    Attempt to find the physical file for a showcase name.
    Applies known typo corrections (e.g. order_surov -> order_suvorov).
    Logs a warning and returns None when not found.
    """
    folder = ASSET_FOLDER.get(country_key, '')
    candidates: list[str] = [showcase_name]

    # Known typo in Excel: order_surov_big1 -> order_suvorov_big1
    if 'surov' in showcase_name and 'suvorov' not in showcase_name:
        candidates.append(showcase_name.replace('surov', 'suvorov'))

    for candidate in candidates:
        path = assets_dir / folder / f"{candidate}.png"
        if path.exists():
            return path

    logger.warning(
        "[medal_showcase] Asset not found for showcase name '%s' "
        "(country='%s', tried %s)",
        showcase_name, country_key, candidates,
    )
    return None


# ---------------------------------------------------------------------------
# Showcase builder
# ---------------------------------------------------------------------------

def build_showcase_data(
    country_key: str,
    earned_showcase_names: set[str],
    coordinates: dict[str, CountryShowcase],
    assets_dir: Path,
    tracker_asset_url_prefix: str = '/api/tracker_assets',
) -> dict:
    """
    Build the JSON-serialisable showcase payload for the frontend.

    Args:
        country_key:            e.g. "ussr_late", "germany"
        earned_showcase_names:  set of showcase names the pilot has earned,
                                e.g. {"iron_cross_2nd_big1", "dfc_big1"}
        coordinates:            parsed Excel data
        assets_dir:             path to CampaignRanksAwards/ root directory
        tracker_asset_url_prefix: URL prefix for tracker assets

    Returns:
        {
            "country_key": "germany",
            "canvas_url":  "/api/tracker_assets/CampaignRanksAwards/Germany/Germany_canvas.png",
            "overlay": {
                "url": "...",
                "x": 32, "y": 0, "w": 2018, "h": 961
            },
            "medals": [
                { "image_url": "...", "x": 141, "y": 134, "w": 363, "h": 503, "name": "iron_cross_2nd_big1" },
                ...
            ]
        }
    """
    showcase = coordinates.get(country_key)
    if showcase is None:
        logger.warning(
            "[medal_showcase] No coordinate data for country key '%s'",
            country_key,
        )
        return {}

    folder = ASSET_FOLDER.get(country_key, '')

    def _make_url(filename: str) -> str:
        return f"{tracker_asset_url_prefix}/CampaignRanksAwards/{folder}/{filename}"

    # Canvas URL
    canvas_file = CANVAS_FILENAME.get(country_key, '')
    canvas_url  = _make_url(canvas_file) if canvas_file else ''

    # Overlay
    overlay_dict: Optional[dict] = None
    if showcase.overlay:
        ovl = showcase.overlay
        overlay_file = OVERLAY_FILENAME.get(country_key, '')
        if overlay_file:
            overlay_dict = {
                'url': _make_url(overlay_file),
                'x': ovl.x,
                'y': ovl.y,
                'w': ovl.w,
                'h': ovl.h,
            }

    # Medals – only earned slots
    placed_medals: list[dict] = []

    for entry in showcase.medals:
        chain = entry.full_chain()  # [base, rep1, rep2, …]  lowest to highest

        # Walk from HIGHEST to LOWEST to find the best earned variant
        chosen: Optional[str] = None
        for candidate in reversed(chain):
            # Normalise for comparison (strip trailing whitespace)
            c = candidate.strip()
            if c in earned_showcase_names:
                chosen = c
                break

        if chosen is None:
            continue  # player hasn't earned this award

        # Verify asset file exists
        asset_path = _resolve_asset_file(chosen, country_key, assets_dir)
        if asset_path is None:
            # Try base entry as fallback
            if chosen != entry.name:
                asset_path = _resolve_asset_file(entry.name, country_key, assets_dir)
            if asset_path is None:
                continue  # no usable image

        placed_medals.append({
            'image_url': _make_url(asset_path.name),
            'x':         entry.x,
            'y':         entry.y,
            'w':         entry.w,
            'h':         entry.h,
            'name':      chosen,
        })

    return {
        'country_key': country_key,
        'canvas_url':  canvas_url,
        'overlay':     overlay_dict,
        'medals':      placed_medals,
    }


# ---------------------------------------------------------------------------
# Convenience: extract earned showcase names from campaign event list
# ---------------------------------------------------------------------------

def earned_showcase_names_from_events(
    events: list[dict],
) -> set[str]:
    """
    Extract the set of showcase image names (e.g. "dfc_big1") from
    campaign events.  Only award-type events contribute.

    USSR early/late note: both early and late images map to the same
    showcase name (the filename stem + _big1), so no era-specific logic
    is needed here.
    """
    result: set[str] = set()
    for ev in events:
        if ev.get('type') != 'award':
            continue
        image = ev.get('image', '')
        name  = award_image_to_showcase_name(image)
        if name:
            result.add(name)
    return result
